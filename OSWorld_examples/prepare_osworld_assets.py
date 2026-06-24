from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "osworld_windows_recommended_tasks_with_links.xlsx"
ASSETS_DIR = BASE_DIR / "assets"
INITIAL_DIR = ASSETS_DIR / "initial"
REFERENCE_DIR = ASSETS_DIR / "reference"
SOURCE_DIR = ASSETS_DIR / "source_json"
EXTERNAL_DIR = ASSETS_DIR / "external_targets"
MANIFEST_PATH = BASE_DIR / "manifest.json"

BACKSLASH = chr(92)
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"


def norm(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def iter_rows_dict(worksheet):
    headers = [norm(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        yield {headers[index]: row[index] for index in range(len(headers))}


def leaf_name(value: str | None, fallback: str) -> str:
    if value:
        normalized = value.replace(BACKSLASH, "/").replace("；", ";")
        first = normalized.split(";")[0].strip()
        return Path(first).name
    return Path(fallback).name


def safe_asset_name(task_id: str, filename: str) -> str:
    invalid = '<>:"/\\|?*'
    table = str.maketrans({char: "_" for char in invalid})
    return f"{task_id}__{filename.translate(table)}"


def ensure_dirs() -> None:
    for directory in [INITIAL_DIR, REFERENCE_DIR, SOURCE_DIR, EXTERNAL_DIR]:
        directory.mkdir(parents=True, exist_ok=True)


def download_file(session: requests.Session, url: str, destination: Path, retries: int = 4) -> None:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            with session.get(url, timeout=120, stream=True) as response:
                response.raise_for_status()
                with destination.open("wb") as handle:
                    for chunk in response.iter_content(chunk_size=1024 * 128):
                        if chunk:
                            handle.write(chunk)
            return
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if destination.exists():
                destination.unlink()
            if attempt < retries:
                time.sleep(min(2 * attempt, 6))
    assert last_error is not None
    raise last_error


def build_manifest() -> dict[str, Any]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    ws_tasks = workbook["Recommended 10"]
    ws_links = workbook["Required file links"]

    tasks: list[dict[str, Any]] = []
    for row in iter_rows_dict(ws_tasks):
        task = {key: norm(value) for key, value in row.items()}
        task["Rank"] = int(row["Rank"])
        tasks.append(task)

    links_by_task: dict[str, list[dict[str, Any]]] = {}
    all_assets: list[dict[str, Any]] = []

    for row in iter_rows_dict(ws_links):
        entry = {key: norm(value) for key, value in row.items()}
        entry["Rank"] = int(row["Rank"])
        task_id = entry["Task ID"]
        role = entry["File role"]
        filename = entry["File / target"]
        default_dest = entry["Default local path / dest"]

        if role == "initial":
            local_name = leaf_name(default_dest, filename.split(" / ")[0])
            asset_rel = f"assets/initial/{safe_asset_name(task_id, local_name)}"
        elif role.startswith("reference"):
            local_name = leaf_name(default_dest, filename.split(" / ")[0])
            asset_rel = f"assets/reference/{safe_asset_name(task_id, local_name)}"
        elif role == "external target":
            local_name = filename.replace("/", "-").replace(" ", "_") + ".url"
            asset_rel = f"assets/external_targets/{safe_asset_name(task_id, local_name)}"
        else:
            raise ValueError(f"Unsupported file role: {role}")

        entry["local_asset"] = asset_rel
        links_by_task.setdefault(task_id, []).append(entry)
        all_assets.append(entry)

    manifest_tasks: list[dict[str, Any]] = []
    for task in tasks:
        task_id = task["Task ID"]
        manifest_tasks.append(
            {
                "rank": task["Rank"],
                "task_id": task_id,
                "source_set": task["Source set"],
                "apps": task["App(s)"],
                "instruction": task["Instruction"],
                "difficulty": task["Difficulty"],
                "windows_suitability": task["Windows suitability"],
                "why_include": task["Why include"],
                "source_file_url": task["Source file URL"],
                "source_json_local": f"assets/source_json/{task['Rank']:02d}_{task_id}.json",
                "reference_gold_urls": task["Reference/gold URL(s)"],
                "external_target_url": task["External target URL"],
                "file_link_notes": task["File/link notes"],
                "assets": links_by_task.get(task_id, []),
            }
        )

    return {
        "generated_from": WORKBOOK_PATH.name,
        "tasks_count": len(tasks),
        "asset_counts": {
            "initial": sum(1 for item in all_assets if item["File role"] == "initial"),
            "reference_gold": sum(1 for item in all_assets if item["File role"] == "reference/gold"),
            "reference_original": sum(1 for item in all_assets if item["File role"] == "reference/original"),
            "external_target": sum(1 for item in all_assets if item["File role"] == "external target"),
            "source_json": len(tasks),
        },
        "tasks": manifest_tasks,
    }


def materialize_assets(manifest: dict[str, Any]) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    downloaded = {
        "source_json": 0,
        "linked_files": 0,
        "external_shortcuts": 0,
    }

    for task in manifest["tasks"]:
        source_url = task["source_file_url"].replace(
            "https://github.com/xlang-ai/OSWorld/blob/main/",
            "https://raw.githubusercontent.com/xlang-ai/OSWorld/main/",
        )
        source_dest = BASE_DIR / task["source_json_local"]
        if not source_dest.exists():
            download_file(session, source_url, source_dest)
            downloaded["source_json"] += 1

        for asset in task["assets"]:
            local_path = BASE_DIR / asset["local_asset"]
            if asset["File role"] == "external target":
                if not local_path.exists():
                    local_path.write_text(
                        "[InternetShortcut]\nURL=" + asset["Download or target URL"] + "\n",
                        encoding="utf-8",
                    )
                    downloaded["external_shortcuts"] += 1
                continue
            if not local_path.exists():
                download_file(session, asset["Download or target URL"], local_path)
                downloaded["linked_files"] += 1

    return downloaded


def main() -> None:
    ensure_dirs()
    manifest = build_manifest()
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=True, indent=2), encoding="utf-8")
    downloaded = materialize_assets(manifest)
    print(
        json.dumps(
            {
                "manifest": str(MANIFEST_PATH.relative_to(BASE_DIR)).replace(BACKSLASH, "/"),
                "downloaded": downloaded,
                "asset_counts": manifest["asset_counts"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
